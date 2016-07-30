#!/usr/bin/env python
# -*- coding: latin-1 -*-

""" Takes one argument, the location of our Jersey Boys spreadsheet """

from openpyxl import load_workbook
import sys

wb = load_workbook("sys.argv[1]")
ws = wb.active

celldict = {}

""" 
Iterates through the seat numbers column, formats a little, using unique sets as a key and all the 
dates they appear as the value pair. Iter_rows needs to be amended for future runs.
"""
for row in ws.iter_rows('D4684:D4879'):
	for cell in row:
		try:
			snums = "".join(str(cell.value).split())
			seats = ws["C" + str(cell.row)].value + " " + snums
			price = ws["E" + str(cell.row)].value
			key = (seats, price)
			
			if key not in celldict.keys():
				celldict[key] = [(ws["A" + str(cell.row)].value, ws["B" + str(cell.row)].value, ws["F" + str(cell.row)].value)]
			else:
				celldict[key].append((ws["A" + str(cell.row)].value, ws["B" + str(cell.row)].value, ws["F" + str(cell.row)].value))
				
		except TypeError:
			pass

""" 
Writes them to jb.txt
"""
with open('jb.txt', 'w') as f:
	for i in celldict.keys():
		try:
			f.write(str(i[0]) + " at " + i[1].encode("utf8") + "\n")
		except AttributeError:
			f.write(str(i[0]) + " at £" + str(i[1]) + "\n")
		for x in celldict[i]:
			date = x[0].strftime('%a %d-%b-%Y')
			f.write(date + " " + str(x[1]) + " " + str(x[2]) + "\n")
		f.write("\n")